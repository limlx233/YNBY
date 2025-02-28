import re
import calendar
import pandas as pd
from io import BytesIO
import streamlit as st
import dataprocess as dp  # 根据实际处理需求 编写的数据处理模块
import concurrent.futures
from datetime import date
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.formatting.rule import DataBarRule, FormulaRule

def add_data_bar_rule(worksheet, start_row, end_row, column, color="c00000"):
    # 范围字符串
    range_str = f'{column}{start_row}:{column}{end_row}'
    # 添加条件格式规则，排除负数
    negative_rule = FormulaRule(formula=[f'AND({column}{start_row}<0)'], stopIfTrue=True)
    worksheet.conditional_formatting.add(range_str, negative_rule)
    # 添加数据条规则
    data_bar_rule = DataBarRule(
        start_type='num', start_value=0,
        end_type='max',
        color=color
    )
    worksheet.conditional_formatting.add(range_str, data_bar_rule)

# 使用缓存来存储处理后的 DataFrame
@st.cache_data
def load_excel_files(uploaded_files):
    dfs = {}
    for uploaded_file in uploaded_files:
        df = pd.read_excel(uploaded_file, header=17)
        dfs[uploaded_file.name] = df
    return dfs

# 使用缓存来存储处理后的数据
@st.cache_data
def process_data(dfs):
    df_lst = []
    for df in dfs.values():
        df = df.dropna(subset=["生产日期", "失效日期"], inplace=False)
        df_lst.append(df)
    df_all = pd.concat(df_lst, axis=0)
    return df_all

# 将 Pandas DataFrame 对象转换为 Excel 文件格式的字节流
@st.cache_resource
def to_excel(df1, df3 ,df2,sheet_name1='物料',sheet_name3='成品',sheet_name2='异常类别定义'):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df2.to_excel(writer, index=False, header=False,sheet_name=sheet_name2)
    df1.to_excel(writer, index=False, sheet_name=sheet_name1)
    df3.to_excel(writer, index=False, sheet_name=sheet_name3)
    # 获取工作簿和工作表
    workbook = writer.book
    worksheet1 = writer.sheets[sheet_name1]
    worksheet2= writer.sheets[sheet_name2]
    worksheet3= writer.sheets[sheet_name3]

        # 确保百分比样式只被创建一次
    if "percentage_style" not in workbook.named_styles:
        percentage_style = NamedStyle(name="percentage_style", number_format='0.00%')
        workbook.add_named_style(percentage_style)
    else:
        percentage_style = workbook.named_styles["percentage_style"]
    # 调用格式设置函数
    set_description_sheet_format(worksheet2, df2)
    set_material_sheet_format(worksheet1, df1, percentage_style)
    set_material_sheet_format(worksheet3, df3, percentage_style)
    # 增加数据条
    add_data_bar_rule(worksheet1,start_row=2, end_row=len(df1) + 1, column='F')
    add_data_bar_rule(worksheet3,start_row=2, end_row=len(df3) + 1, column='F')
    writer.close()
    processed_data = output.getvalue()
    return processed_data

def set_material_sheet_format(worksheet, df1, percentage_style):
    # 设置列宽
    for idx, column in enumerate(worksheet.columns, start=1):
        column_letter = column[0].column_letter
        if idx == 1:
            worksheet.column_dimensions[column_letter].width = 14
        elif idx == 6:
            worksheet.column_dimensions[column_letter].width = 22
        elif idx == 8:
            worksheet.column_dimensions[column_letter].width = 30
        else:
            worksheet.column_dimensions[column_letter].width = 16
    # 应用百分比格式到指定列
    column_name = '%(剩余效期/总效期)'
    column_index = df1.columns.get_loc(column_name) + 1  # +1 是因为 Excel 列索引从 1 开始
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            cell.style = percentage_style
    # 定义将 RGB 颜色值转换为十六进制颜色代码的函数
    def rgb_to_hex(r, g, b):
        return '{:02x}{:02x}{:02x}'.format(r, g, b)
    # 定义不同列所需的颜色
    colors = {
        1: rgb_to_hex(226, 107, 10),
        2: rgb_to_hex(49, 134, 155),
        3: rgb_to_hex(226, 107, 10),
        4: rgb_to_hex(226, 107, 10),
        5: rgb_to_hex(226, 107, 10),
        6: rgb_to_hex(118, 147, 60),
        7: rgb_to_hex(118, 147, 60)
    }
    # 默认颜色
    default_color = "346c9c"
    # 遍历第一行的每个单元格
    for col_index, cell in enumerate(worksheet[1], start=1):
        # 设置字体为加粗，颜色为白色
        cell.font = Font(bold=True, color="FFFFFF")
        # 根据列索引获取相应的颜色，如果没有指定则使用默认颜色
        fill_color = colors.get(col_index, default_color)
        # 设置填充颜色
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        # 设置单元格内容居中对齐
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if cell.column == 8 :  # G列是第7列 or cell.column == 14
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def set_description_sheet_format(worksheet, df2):
    # 设置列宽
    worksheet.column_dimensions['A'].width = 22
    worksheet.column_dimensions['B'].width = 108
    # 设置行高
    for row in range(1, 12):  # 1~11行行高设置为36
        worksheet.row_dimensions[row].height = 36
    # 设置标题样式
    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 设置剩余行的标题样式
    for row in range(2, worksheet.max_row + 1):  # 从第二行开始
        for cell in worksheet[row]:
            cell.font = Font(size=14, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    # 合并A1和B1
    worksheet.merge_cells('A1:B1')
    # 设置合并单元格的对齐方式
    worksheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    # 保证样式设置不会被合并操作覆盖
    for row in worksheet['A1:B1']:
        for cell in row:
            cell.font = Font(bold=True, size=16)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # 设置A2:A8单元格的样式
    for row in worksheet['A2:A11']:
        for cell in row:
            cell.font = Font(size=14, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    # 设置C2:C8单元格的样式
    for row in worksheet['C2:C11']:
        for cell in row:
            cell.font = Font(size=14, color="000000", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

# 页面设置
st.set_page_config(page_title="数据处理工具", page_icon=":material/home:", layout='centered')

with st.container(border=True):
    st.header('每月物料库存数据处理', divider="rainbow")
    st.subheader('月末日期选择', divider='grey')
    def get_last_day_of_previous_month():
        today = date.today()
        if today.month == 1:
            year = today.year - 1
            month = 12
        else:
            year = today.year
            month = today.month - 1
        _, last_day = calendar.monthrange(year, month)
        return date(year, month, last_day)
    
    # 默认日期为上月最后一天
    default_date = get_last_day_of_previous_month()
    # 日期输入控件
    date_value = st.date_input(label="请选择日期,(默认为上月的最后一天)", value=default_date)
    # st.write(date_value)
    st.subheader('数据文件上传', divider='grey')
    # 多文件上传
    uploaded_files = st.file_uploader(label="请选择Excel文件(.xlsx格式)上传", accept_multiple_files=True, type=["xlsx"])
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button(label="数据处理", type="primary", key="data_process"):
            if uploaded_files:
                # 读取文件并缓存
                dfs = load_excel_files(uploaded_files)
                columns_to_keep = st.secrets["warehouses"]["columns_to_keep"]
                # 并行处理数据
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    df_lst = list(executor.map(lambda df: df.dropna(subset=["生产日期", "失效日期"], inplace=False)[columns_to_keep], dfs.values()))
                df_all = pd.concat(df_lst, axis=0)
                # 定义正则表达式，匹配字母和数字字符
                pattern = re.compile(r'^([^:]+):')
                # 使用正则表达式提取仓库代码并创建新列
                df_all['仓库代码'] = df_all['仓库'].apply(lambda x: pattern.match(x).group(1) if pattern.match(x) else None)
                # 数据处理--物料
                wl = st.secrets["warehouses"]["wl"]
                cp_wx = st.secrets["warehouses"]["cp_wx"]
                df_wl1 = dp.warehouse_filtering(df_all,wl) # 筛选仓库
                df_wl1 = dp.calculate_expiry(df_wl1, date_value) # 效期计算
                df_wl1 = dp.expiry_classification(df_wl1) # 效期类别
                df_wl1 = dp.receive_classification(df_wl1,date_value) # 领用时间分类
                df_wl1 = dp.storage_days_classification(df_wl1,cp_wx) # 在库时间分类
                df_wl1 = df_wl1[~((df_wl1['效期类别'] == '' ) & (df_wl1['90天内无领用'] == '' ) & (df_wl1['异常在库天数'] == ''))]
                df_wl1 = dp.classify_items(df_wl1)
                df_wl1['处理方案'] =''
                df_wl1 = dp.sort_and_filter(df_wl1)
                # 数据处理-成品
                cp = st.secrets["warehouses"]["cp_warehouses"]
                cp_filter = st.secrets["warehouses"]["cp"]
                df_cp1 = dp.cp_warehouse_filtering(df_all,cp,cp_filter)
                df_cp1 = dp.calculate_expiry(df_cp1, date_value) # 效期计算
                df_cp1 = dp.expiry_classification(df_cp1) # 效期类别
                df_cp1 = dp.receive_classification(df_cp1,date_value) # 领用时间分类
                df_cp1 = dp.storage_days_classification(df_cp1,cp_wx) # 在库时间分类
                df_cp1 = df_cp1[~((df_cp1['效期类别'] == '' ) & (df_cp1['90天内无领用'] == '' ) & (df_cp1['异常在库天数'] == ''))]
                df_cp1 = dp.classify_items(df_cp1)
                df_cp1['处理方案'] =''
                # 重新对列排序
                df_cp1 = dp.sort_and_filter(df_cp1)
                # st.write(df_all.shape)
                # st.data_editor(df_wl1)
                # 生成 Excel 文件
                df2 = dp.generate_description_df()
                excel_file = to_excel(df_wl1,df_cp1,df2)
                st.session_state.excel_file = excel_file
            else:
                st.info("请先上传数据文件!")
    with col2:
        if 'excel_file' in st.session_state:
            st.download_button(
                label="下载文件",
                data=st.session_state.excel_file,
                type="primary",
                file_name="月末库存呆滞情况.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("请先上传数据并进行数据处理。")